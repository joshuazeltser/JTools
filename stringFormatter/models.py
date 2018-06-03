import openpyxl as openpyxl
from django.db import models
from random import shuffle


class Text(models.Model):
    text_id = models.IntegerField(default=0, unique=True)
    text = models.CharField(max_length=1000)
    result = models.CharField(default=0, max_length=100)

    def __str__(self):
        return self.text

    def to_upper_case(string):
        return string.upper()

    def to_lower_case(string):
        return string.lower()

    def word_count(string):
        words = string.split()

        return len(words)

    def char_count(string):
        return len(string)

    def split_string_by(string, char):
        return string.split(char)

    def word_occurence_count(string):
        words = string.split()

        dict = {}

        for x in range(0, len(words)):
            dict[words[x]] = words.count(words[x])

        string = ''
        count = 1
        for key in dict:
            string += key + ': ' + str(dict[key])
            if count < len(dict):
                string += ', '

            count += 1

        return string


class RandomList(models.Model):
    list = []
    wb = openpyxl.load_workbook('randomised_list.xlsx')
    ws = wb.active
    count = 1

    def populate_list(self, l):
        self.list = l[:]

    def shuffle_list(self):
        shuffle(self.list)

    def write_to_spreadhseet(self):
        line = 1
        for td in range(len(self.list)):
            self.ws.cell(row=line, column=1, value=self.list[td])
            line += 1
        self.wb.save('randomised_list.xlsx')

    def create_new_worksheet(self):
        self.wb.create_sheet('Sheet ' + str(self.count))
        self.ws = self.wb['Sheet ' + str(self.count)]
        self.count += 1
        self.wb.save('randomised_list.xlsx')

    def random_sublists_to_excel(self, num):
        self.shuffle_list()

        chunks = [self.list[x:x+num] for x in range(0, len(self.list), num)]

        column = 1
        for group in range(len(chunks)):
            self.ws.cell(row=1, column=column, value='Group ' + str(column))
            line = 2
            for td in range(len(chunks[group])):
                self.ws.cell(row=line, column=column, value=chunks[group][td])
                line += 1

            column += 1
        self.wb.save('randomised_list.xlsx')


    def random_sublists_to_string(self, num):

        result = ""

        self.shuffle_list()

        chunks = [self.list[x:x+num] for x in range(0, len(self.list), num)]

        column = 1
        for group in range(len(chunks)):
            result += 'Group ' + str(column) + '/n'
            line = 2
            for td in range(len(chunks[group])):
                result += chunks[group][td]
                line += 1

            column += 1